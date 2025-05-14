from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
import smtplib
from email.message import EmailMessage
import os
from openpyxl import Workbook, load_workbook
import gspread 
from oauth2client.service_account import ServiceAccountCredentials

app = Flask(__name__, static_folder='statics', template_folder='templates')
CORS(app)

EMAIL_ADDRESS = "swarajbehera.8673@gmail.com"
EMAIL_PASSWORD = "vkwdrwkrsjfjadpc"

SHEET_NAME = 'Client Data'  # Name of your Google Sheet
CREDENTIALS_FILE = 'quixotic-tesla-459406-u5-b0653b0e1c15.json'  # Path to your service account credentials

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/send-email', methods=['POST'])
def send_email():
    data = request.get_json()

    name = data.get('name')
    address = data.get('address')
    whatsapp = data.get('whatsapp')
    email = data.get('email')

    # # === EXCEL LOGIC START ===
    # excel_file = "client_data.xlsx"
    # if not os.path.exists(excel_file):
    #     wb = Workbook()
    #     ws = wb.active
    #     ws.append(["Name", "Address", "WhatsApp", "Email"])
    # else:
    #     wb = load_workbook(excel_file)
    #     ws = wb.active

    # ws.append([name, address, whatsapp, email])
    # wb.save(excel_file)
    # # === EXCEL LOGIC END ===

    # === GOOGLE SHEETS LOGIC START ===
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, scope)
    client = gspread.authorize(creds)

    sheet = client.open("Client Data").sheet1  # Replace with your sheet name

    # Append data to the sheet
    sheet.append_row([name, address, whatsapp, email])

    # Get the sheet URL
    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet.spreadsheet.id}/edit"
    # === GOOGLE SHEETS LOGIC END ===

    msg = EmailMessage()
    msg['Subject'] = 'New Contact Form Submission'
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = EMAIL_ADDRESS
    msg.set_content(f"Name: {name}\nAddress: {address}\nWhatsApp: {whatsapp}\nEmail: {email}\nGoogle Sheet Link: {sheet_url}")

    # # Attach the Excel file
    # with open(excel_file, 'rb') as f:
    #     msg.add_attachment(f.read(), maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=excel_file)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)

        return jsonify({"message": "Email and Excel attachment sent successfully!"}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "Failed to send email."}), 500

if __name__ == '__main__':
    app.run(debug=True)
